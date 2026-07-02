import streamlit as st
import fitz
import pytesseract
import pandas as pd
from PIL import Image
import io
import os
import re
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    Border,
    Side,
    Font,
    PatternFill,
    Alignment
)
from langchain_ollama import OllamaLLM
from langchain_core.prompts import PromptTemplate
# ============================================================
# CONFIG
# ============================================================
BASE_PATH = r"C:\AIFoundaryWorkshop"

INPUT_FOLDER = os.path.join(BASE_PATH, "Input")

OUTPUT_FOLDER = os.path.join(
    BASE_PATH,
    "Translated_Output"
)

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

TESS_PATH = os.path.join(
    BASE_PATH,
    "Tesseract-OCR",
    "tesseract.exe"
)

if os.path.exists(TESS_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESS_PATH


# ============================================================
# UI
# ============================================================
st.set_page_config(
    page_title="Opella AI Japanese to English Translator",
    layout="wide"
)

st.markdown(
    """
<style>

.stApp {
    background-color: #042B0B;
}

.stButton>button {
    background-color:#6D7F6E;
    color:white;
    border-radius:12px;
    font-weight:bold;
}

h1,h2,h3 {
    color:#042B0B;
}

</style>
""",
    unsafe_allow_html=True
)

logo = os.path.join(
    BASE_PATH,
    "Opella_Logo.png"
)

if os.path.exists(logo):
    st.image(logo, width=180)


# ============================================================
# OCR
# ============================================================
def ocr_with_pos(img):

    df = pytesseract.image_to_data(
        img,
        lang="jpn+jpn_vert",
        output_type=pytesseract.Output.DATAFRAME
    )

    df = df.dropna()
    df = df[df.text.str.strip() != ""]

    return df


# ============================================================
# BUILD TABLE
# ============================================================
def build_table(df):

    rows = []

    df["row_id"] = df["top"] // 20

    for _, row in df.groupby("row_id"):

        words = sorted(
            zip(
                row["left"],
                row["text"]
            )
        )

        rows.append(
            [w[1] for w in words]
        )

    if not rows:
        return pd.DataFrame()

    max_cols = max(len(r) for r in rows)

    rows = [
        r + [""] * (max_cols - len(r))
        for r in rows
    ]

    return pd.DataFrame(rows)


# ============================================================
# CLEANING
# ============================================================
def clean_translation(text):

    text = re.sub(r"\*+", "", text)

    text = re.sub(
        r"(?i)here is the translation.*",
        "",
        text
    )

    text = re.sub(
        r"(?i)please note.*",
        "",
        text
    )

    return text.strip()


# ============================================================
# TEXT TRANSLATION
# ============================================================
def translate_text(text, model):

    llm = OllamaLLM(model=model)

    prompt = PromptTemplate.from_template(
        """
You are a professional Japanese business translator.

Translate the document into fluent business English.

Requirements:

- Translate ALL Japanese text.
- Keep headings.
- Preserve names.
- Preserve addresses.
- Preserve dates.
- Preserve phone numbers.
- Preserve permit numbers.
- Do NOT summarize.
- Do NOT explain.
- Do NOT add notes.
- Do NOT say "Here is the translation".
- Return only the translated document.

DOCUMENT:

{text}
"""
    )

    chain = prompt | llm

    chunks = [
        text[i:i + 2500]
        for i in range(
            0,
            len(text),
            2500
        )
    ]

    translated = []

    for chunk in chunks:

        translated.append(
            chain.invoke(
                {"text": chunk}
            )
        )

    return clean_translation(
        "\n".join(translated)
    )


# ============================================================
# TABLE EXTRACTION
# ============================================================

# ============================================================
# TABLE EXTRACTION HELPERS
# ============================================================

def contains_japanese(text):

    return bool(
        re.search(
            r'[\u3040-\u30ff\u4e00-\u9fff]',
            str(text)
        )
    )

def is_noise(text):

    text = str(text)

    # very short garbage
    if len(text.strip()) < 3:
        return True

    # too many symbols
    symbol_count = len(
        re.findall(
            r'[^A-Za-z0-9\u3040-\u30ff\u4e00-\u9fff\s]',
            text
        )
    )

    if symbol_count > len(text) * 0.4:
        return True

    return False


# ============================================================
# TABLE EXTRACTION
# ============================================================

def structure_table(df, model):

    rows = []

    for _, row in df.iterrows():

        vals = [
            str(x).strip()
            for x in row
            if str(x).strip()
        ]

        if vals:
            rows.append(" ".join(vals))

    ocr_text = "\n".join(rows)

    llm = OllamaLLM(model=model)

    prompt = PromptTemplate.from_template(
        """
Extract business information from this Japanese document.

Translate everything into English.

Return ONLY VALID JSON.

Example:

[
    {{
        "Field":"Permission Number",
        "Value":"0108241929"
    }},
    {{
        "Field":"Company Name",
        "Value":"Moz Pharmaceutical Co., Ltd."
    }},
    {{
        "Field":"Address",
        "Value":"1-8 Yurakucho, Tokyo"
    }},
    {{
        "Field":"Phone",
        "Value":"03-5844-0252"
    }}
]

TEXT:

{text}
"""
    )

    chain = prompt | llm

    response = chain.invoke(
        {"text": ocr_text}
    )

    try:

        start = response.find("[")
        end = response.rfind("]")

        if start == -1 or end == -1:
            raise ValueError(
                "JSON not found"
            )

        json_text = response[
            start:end + 1
        ]

        records = json.loads(
            json_text
        )

        df_out = pd.DataFrame(
            records
        )

        if "Field" not in df_out.columns:
            return pd.DataFrame(
                columns=[
                    "Field",
                    "Value"
                ]
            )

        if "Value" not in df_out.columns:
            return pd.DataFrame(
                columns=[
                    "Field",
                    "Value"
                ]
            )

        df_out = (
            df_out
            .drop_duplicates()
            .reset_index(drop=True)
        )

        # =====================================
        # Translate any remaining Japanese
        # =====================================

        for idx, row in df_out.iterrows():

            field = str(row["Field"])
            value = str(row["Value"])

            # Translate Field
            if contains_japanese(field):

                try:

                    translated_field = llm.invoke(
                        f"""
        Translate the following Japanese business field into English.

        Return only the translation.

        Text:
        {field}
        """
                    )

                    df_out.at[idx, "Field"] = translated_field.strip()

                except Exception:
                    pass

            # Translate Value
            if contains_japanese(value) and not is_noise(value):

                try:

                    translated_value = llm.invoke(
                        f"""
                    Translate the following Japanese text into English.

                    Rules:
                    - Return ONLY the translation.
                    - If the text is unreadable, corrupted OCR, or not valid Japanese,
                    return exactly: SKIP

                    Text:
                    {value}
                    """
                    ).strip()

                    bad_responses = [
                        "Unfortunately",
                        "I have to inform you",
                        "not a valid Japanese",
                        "cannot translate",
                        "SKIP"
                    ]

                    if not any(
                        x.lower() in translated_value.lower()
                        for x in bad_responses
                    ):
                        df_out.at[idx, "Value"] = translated_value
                except Exception:
                    pass

        return df_out
# =====================================


    except Exception:

        return pd.DataFrame(
            columns=[
                "Field",
                "Value"
            ]
        )


# ============================================================
# FORMAT EXCEL
# ============================================================
def format_excel(path):

    wb = load_workbook(path)

    ws = wb.active

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            size=12
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

        cell.border = border

    for row in ws.iter_rows():

        for cell in row:

            cell.border = border

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 100

    wb.save(path)


# ============================================================
# PDF PROCESSING
# ============================================================
def process_pdf(
    path,
    model,
    progress_bar=None
):

    doc = fitz.open(path)

    text = ""
    tables = []

    total_pages = len(doc)

    for page_num, page in enumerate(doc):

        if progress_bar:

            progress_bar.progress(
                (page_num + 1)
                / total_pages
            )

        pix = page.get_pixmap(
            matrix=fitz.Matrix(5, 5),
            alpha=False
        )

        img = Image.open(
            io.BytesIO(
                pix.tobytes()
            )
        )

        df_pos = ocr_with_pos(img)

        tbl = build_table(df_pos)

        if not tbl.empty:
            tables.append(tbl)

        text += (
            " ".join(
                df_pos["text"].tolist()
            )
            + "\n"
        )

    translated_text = translate_text(
        text,
        model
    )

    
    original_filename = os.path.splitext(
        os.path.basename(path)
    )[0]

    filename = translate_filename(
        original_filename,
        model
    )


    txt_path = os.path.join(
        OUTPUT_FOLDER,
        filename + "_EN.txt"
    )

    with open(
        txt_path,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(
            translated_text
        )

    if tables:

        combined = pd.concat(
            tables,
            ignore_index=True
        )

        structured_df = structure_table(
            combined,
            model
        )

    else:

        structured_df = pd.DataFrame(
            columns=[
                "Field",
                "Value"
            ]
        )

    excel_path = os.path.join(
        OUTPUT_FOLDER,
        filename + ".xlsx"
    )

    structured_df.to_excel(
        excel_path,
        index=False
    )

    format_excel(excel_path)

    return (
        translated_text,
        structured_df,
        excel_path
    )

#---------translate file name-------------------------------------------------

def translate_filename(filename, model):

    try:

        llm = OllamaLLM(model=model)

        prompt = f"""
Translate this Japanese business name into English.

Rules:
- Return only the English name.
- No explanation.
- Replace spaces with underscores.
- Use only letters, numbers and underscores.

Text:
{filename}
"""

        result = llm.invoke(prompt)

        result = result.strip()

        result = re.sub(r'[^A-Za-z0-9_ ]', '', result)
        result = result.replace(" ", "_")

        if result:
            return result

    except:
        pass

    return filename

# ============================================================
# BATCH PROCESSING
# ============================================================
def process_folder(
    folder_path,
    model
):

    files = [
        f
        for f in os.listdir(
            folder_path
        )
        if f.lower().endswith(".pdf")
    ]

    if not files:

        st.warning(
            "No PDF files found."
        )
        return

    overall = st.progress(0)

    status = st.empty()

    total = len(files)

    for i, file_name in enumerate(files):

        status.info(
            f"Processing {file_name}"
        )

        page_bar = st.progress(0)

        process_pdf(
            os.path.join(
                folder_path,
                file_name
            ),
            model,
            page_bar
        )

        overall.progress(
            (i + 1) / total
        )

    status.success(
        f"Completed processing {total} files."
    )


# ============================================================
# UI
# ============================================================
def main():

    st.title(
        "🚀 Opella Enterprise Translator"
    )

    model = st.selectbox(
        "Model",
        ["llama3", "mistral"]
    )

    st.subheader(
        "📄 Single PDF"
    )

    file = st.file_uploader(
        "Upload PDF",
        type=["pdf"]
    )

    if file and st.button(
        "Process File 🚀"
    ):

        progress = st.progress(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        base_name = os.path.splitext(file.name)[0]

        temp_path = os.path.join(
            BASE_PATH,
            f"{base_name}_{timestamp}.pdf"
        )

        with open(
            temp_path,
            "wb"
        ) as f:
            f.write(file.read())

        text, df, excel = process_pdf(
            temp_path,
            model,
            progress
        )

        st.success("✅ Done")

        st.info(
            f"Output Folder:\n{OUTPUT_FOLDER}"
        )

        st.subheader(
            "Translated Document"
        )

        st.text_area(
            "",
            text,
            height=300
        )

        st.subheader(
            "Structured Table"
        )

        st.dataframe(df)

        with open(
            excel,
            "rb"
        ) as f:

            st.download_button(
                "Download Excel",
                f,
                file_name=os.path.basename(excel)
            )

    st.markdown("---")

    st.subheader(
        "📂 Batch Processing"
    )

    folder_path = st.text_input(
        "Folder Path",
        value=INPUT_FOLDER
    )

    if st.button(
        "Process Folder 🚀"
    ):

        if os.path.exists(
            folder_path
        ):

            process_folder(
                folder_path,
                model
            )

        else:

            st.error(
                "Folder path not found."
            )


if __name__ == "__main__":
    main()
